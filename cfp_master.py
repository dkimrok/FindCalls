# -*- coding: utf-8 -*-
"""
CFP 마스터 통합 + 신규 감지(diff) 스크립트
============================================
5개 소스의 크롤링 결과를 하나의 마스터 시트로 병합하고, 재실행 시
지난 실행 이후 "새로 등장한" CFP만 골라 보여준다.

입력 (같은 폴더에 있는 파일만 자동 인식; 없으면 건너뜀):
  - cfplist_all.csv        (cfplist.com)
  - sciencedirect_cfps.csv (ScienceDirect — 원시 cardText 자동 파싱)
  - tandf_cfps_v2.csv      (Taylor & Francis API)
  - sage_cfps.csv          (SAGE v7)
  - watchlist_cfps.csv     (INFORMS/OUP/Cambridge 워치리스트)

출력:
  - CFP_master.xlsx  : [신규] [진행중_관심] [전체] 시트
  - cfp_snapshot.json: 지금까지 본 CFP 키 + 최초 관측일 (diff 기준,
                       삭제하지 말 것)

사용법:
  1) 각 크롤러를 돌려 CSV들을 갱신
  2) python cfp_master.py
  3) 콘솔의 "신규 N건"과 CFP_master.xlsx의 [신규] 시트 확인
"""
import json
import re
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TODAY = pd.Timestamp(date.today())
SNAPSHOT = Path("cfp_snapshot.json")

# ── 관심도 태깅 규칙 (필요시 수정) ──────────────────────────────────────
TARGET_JOURNALS = re.compile(
    r"(European Journal of Operational Research|\bOmega\b|"
    r"Computers & Operations Research|Operations Research|Management Science|"
    r"Mathematics of Operations Research|Transportation Science|"
    r"Decision Analysis|INFORMS Journal|M&SOM|Manufacturing & Service|"
    r"Information Systems Research|Engineering Optimization|"
    r"International Journal of Production Research|IISE|"
    r"Energy Policy|Research Policy|Technological Forecasting|Technovation|"
    r"Defen[cs]e|Security|Strategic Studies|Military|Simulation|"
    r"Reliability Engineering|Decision Support|Expert Systems|"
    r"Transportation Research|Technology in Society|"
    r"Energy Research & Social Science|Applied Energy|Applied Economics|"
    r"Science and Public Policy|Industrial and Corporate Change|"
    r"Public Policy|Data & Policy|Korea)", re.I)
TOPIC_KEYWORDS = re.compile(
    r"(defen[cs]e|military|security|war\b|UAV|drone|unmanned|kill.?chain|"
    r"semiconductor|chip\b|supply chain|dual.use|geopolit|deterren|nuclear|"
    r"AI governance|artificial intelligence|generative AI|LLM|"
    r"large language|machine learning|reinforcement learning|"
    r"game.theor|decision.mak|optimi[sz]|operations research|"
    r"vehicle routing|routing|scheduling|quantum|"
    r"energy (security|transition|polic)|critical infrastructure|"
    r"Korea|technology transfer|innovation polic|rare earth|"
    r"critical mineral)", re.I)

COLS = ["출처", "저널/주최", "제목", "초록마감", "원고마감", "마감원문",
        "상태", "관심", "최초관측", "URL"]


def norm_key(url, journal, title):
    if isinstance(url, str) and url.startswith("http"):
        u = url.split("?")[0].split("#")[0].rstrip("/").lower()
        return u
    return (str(journal).strip().lower() + "||" +
            str(title).strip().lower())[:300]


def to_dt(x):
    return pd.to_datetime(x, errors="coerce")


def status_of(ms_dt, ab_dt):
    d = ms_dt if pd.notna(ms_dt) else ab_dt
    if pd.isna(d):
        return "미상"
    return "진행중" if d >= TODAY else "마감"


def interest_of(journal, title):
    j = TARGET_JOURNALS.search(str(journal) or "")
    t = TOPIC_KEYWORDS.search(str(title) or "")
    if j and t:
        return "★★"
    if j or t:
        return "★"
    return ""


def load_cfplist():
    p = Path("cfplist_all.csv")
    if not p.exists():
        return []
    df = pd.read_csv(p, encoding="utf-8-sig")
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "출처": "cfplist", "저널/주최": "",
            "제목": r.get("title", ""),
            "초록마감": to_dt(r.get("abstract_deadline")),
            "원고마감": pd.NaT,
            "마감원문": str(r.get("abstract_deadline") or ""),
            "URL": r.get("url", ""),
        })
    return rows


def load_sciencedirect():
    p = Path("sciencedirect_cfps.csv")
    if not p.exists():
        return []
    df = pd.read_csv(p, encoding="utf-8-sig")
    df = df[df["href"].astype(str).str.contains("/special-issue/", na=False)]
    rows = []
    for _, r in df.iterrows():
        t = str(r.get("cardText") or "")
        lines = [l.strip().replace("\xa0", " ")
                 for l in t.split("\n") if l.strip()]
        title = lines[0] if lines else ""
        jline = next((l for l in lines
                      if "Impact Factor" in l or "CiteScore" in l), "")
        journal = jline.split("\u2022")[0].strip() if jline else ""
        m_dl = re.search(r"Submission deadline:\s*(.+)", t)
        raw = m_dl.group(1).strip() if m_dl else ""
        if not journal:
            dl_idx = next((i for i, l in enumerate(lines)
                           if l.startswith("Submission deadline")), None)
            if dl_idx and dl_idx >= 1 and not lines[dl_idx - 1].lower()\
                    .startswith("guest editor") and lines[dl_idx - 1] != title:
                journal = lines[dl_idx - 1]
        rows.append({
            "출처": "ScienceDirect", "저널/주최": journal, "제목": title,
            "초록마감": pd.NaT,
            "원고마감": pd.to_datetime(raw, format="%d %B %Y",
                                       errors="coerce"),
            "마감원문": raw, "URL": r.get("href", ""),
        })
    return rows


def load_tandf():
    p = Path("tandf_cfps_v2.csv")
    if not p.exists():
        return []
    df = pd.read_csv(p, encoding="utf-8-sig")
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "출처": "T&F", "저널/주최": r.get("journal", ""),
            "제목": r.get("title", ""),
            "초록마감": to_dt(r.get("abstract_deadline_dt")
                              or r.get("abstract_deadline")),
            "원고마감": to_dt(r.get("manuscript_deadline_dt")
                              or r.get("manuscript_deadline")),
            "마감원문": str(r.get("manuscript_deadline") or ""),
            "URL": r.get("cfp_page", ""),
        })
    return rows


def load_generic(path, source_label, journal_col="journal",
                 title_col="si_title"):
    p = Path(path)
    if not p.exists():
        return []
    df = pd.read_csv(p, encoding="utf-8-sig")
    rows = []
    for _, r in df.iterrows():
        pub = r.get("publisher")
        journal = r.get(journal_col, "")
        label = f"{pub}" if isinstance(pub, str) and pub else source_label
        rows.append({
            "출처": label, "저널/주최": journal,
            "제목": r.get(title_col, ""),
            "초록마감": pd.NaT,
            "원고마감": to_dt(r.get("deadline_dt")),
            "마감원문": str(r.get("deadline_raw") or ""),
            "URL": r.get("url", ""),
        })
    return rows


def build_master():
    rows = (load_cfplist() + load_sciencedirect() + load_tandf()
            + load_generic("sage_cfps.csv", "SAGE")
            + load_generic("watchlist_cfps.csv", "Watchlist"))
    if not rows:
        raise SystemExit("입력 CSV가 하나도 없습니다. 크롤러 출력과 같은 "
                         "폴더에서 실행하세요.")
    df = pd.DataFrame(rows)
    df["제목"] = df["제목"].astype(str).str.strip()
    df = df[df["제목"].str.len() >= 8]
    df["키"] = [norm_key(u, j, t) for u, j, t in
                zip(df["URL"], df["저널/주최"], df["제목"])]
    df = df.drop_duplicates(subset="키")
    df["상태"] = [status_of(m, a) for m, a in
                  zip(df["원고마감"], df["초록마감"])]
    df["관심"] = [interest_of(j, t) for j, t in
                  zip(df["저널/주최"], df["제목"])]
    return df


def apply_snapshot(df):
    seen = {}
    if SNAPSHOT.exists():
        seen = json.loads(SNAPSHOT.read_text(encoding="utf-8"))
    today_s = str(date.today())
    first_seen, is_new = [], []
    for k in df["키"]:
        if k in seen:
            first_seen.append(seen[k])
            is_new.append(False)
        else:
            seen[k] = today_s
            first_seen.append(today_s)
            is_new.append(True)
    df["최초관측"] = first_seen
    df["신규"] = is_new
    SNAPSHOT.write_text(json.dumps(seen, ensure_ascii=False),
                        encoding="utf-8")
    return df


def write_xlsx(df):
    wb = Workbook()
    hf = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    bf = Font(name="Arial", size=10)
    fill = PatternFill("solid", fgColor="1F4E79")

    def sheet(ws, d):
        for c, h in enumerate(COLS, 1):
            cell = ws.cell(1, c, h)
            cell.font = hf
            cell.fill = fill
        for r, (_, row) in enumerate(d.iterrows(), 2):
            vals = [row["출처"], row["저널/주최"], row["제목"],
                    "" if pd.isna(row["초록마감"])
                    else row["초록마감"].date().isoformat(),
                    "" if pd.isna(row["원고마감"])
                    else row["원고마감"].date().isoformat(),
                    row["마감원문"], row["상태"], row["관심"],
                    row["최초관측"], row["URL"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.font = bf
                cell.alignment = Alignment(
                    vertical="top", wrap_text=(c == 3))
        for i, w in enumerate([13, 30, 58, 11, 11, 22, 7, 6, 11, 42], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        if len(d):
            ws.auto_filter.ref = f"A1:J{len(d) + 1}"

    new_df = df[df["신규"]].sort_values(["관심", "원고마감"],
                                        ascending=[False, True])
    ws1 = wb.active
    ws1.title = f"신규({len(new_df)})"
    sheet(ws1, new_df)

    hot = df[(df["상태"] == "진행중") & (df["관심"] != "")]\
        .sort_values(["관심", "원고마감"], ascending=[False, True])
    ws2 = wb.create_sheet(f"진행중_관심({len(hot)})")
    sheet(ws2, hot)

    all_df = df.sort_values(["상태", "원고마감"])
    ws3 = wb.create_sheet(f"전체({len(all_df)})")
    sheet(ws3, all_df)

    wb.save("CFP_master.xlsx")


def main():
    df = build_master()
    df = apply_snapshot(df)
    write_xlsx(df)
    print("소스별 건수:")
    print(df["출처"].value_counts().to_string())
    print(f"\n전체 {len(df)}건 | 진행중 {(df['상태'] == '진행중').sum()}건 | "
          f"진행중·관심 {((df['상태'] == '진행중') & (df['관심'] != '')).sum()}건")
    print(f"이번 실행 신규: {df['신규'].sum()}건 → CFP_master.xlsx [신규] 시트")


if __name__ == "__main__":
    main()
